"""
27.15
Скласти програму, яка працює в оточенні веб-сервера, для
конвертування валют. Поточні курси валют містяться у файлі Excel на
сервері у форматі: <код валюти 1> <код валюти 2> <курс>. 

Код валюти – це рядок з 3 символів, наприклад UAH, USD< EUR тощо.
"""

from wsgiref.simple_server import make_server
import urllib.parse
import os

# Try to import openpyxl for Excel support
try:
    import openpyxl
    EXCEL_SUPPORT = True
except ImportError:
    EXCEL_SUPPORT = False
    print("Warning: openpyxl not installed. Run: pip install openpyxl")


def load_exchange_rates(filename="exchange_rates.xlsx"):
    """Load exchange rates from Excel file."""
    rates = {}
    currencies = set()
    
    if not EXCEL_SUPPORT:
        # Default rates if openpyxl is not installed
        default_rates = [
            ("UAH", "USD", 41.50),
            ("UAH", "EUR", 43.20),
            ("USD", "EUR", 1.04),
            ("USD", "UAH", 0.024),
            ("EUR", "UAH", 0.023),
            ("EUR", "USD", 0.96),
        ]
        for c1, c2, rate in default_rates:
            rates[(c1, c2)] = rate
            currencies.add(c1)
            currencies.add(c2)
        return rates, sorted(currencies)
    
    filepath = os.path.join(os.path.dirname(__file__), filename)
    
    if not os.path.exists(filepath):
        # Create default Excel file if it doesn't exist
        create_default_excel(filepath)
    
    try:
        workbook = openpyxl.load_workbook(filepath)
        sheet = workbook.active
        
        # Skip header row, read data
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] and row[1] and row[2]:
                currency1 = str(row[0]).strip().upper()
                currency2 = str(row[1]).strip().upper()
                rate = float(row[2])
                rates[(currency1, currency2)] = rate
                currencies.add(currency1)
                currencies.add(currency2)
    except Exception as e:
        print(f"Error loading Excel file: {e}")
    
    return rates, sorted(currencies)


def create_default_excel(filepath):
    """Create a default Excel file with sample exchange rates."""
    if not EXCEL_SUPPORT:
        return
    
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Exchange Rates"
    
    # Header row
    sheet['A1'] = "Currency1"
    sheet['B1'] = "Currency2"
    sheet['C1'] = "Rate"
    
    # Sample data
    data = [
        ("UAH", "USD", 0.024),
        ("UAH", "EUR", 0.023),
        ("USD", "UAH", 41.50),
        ("USD", "EUR", 0.92),
        ("EUR", "UAH", 43.20),
        ("EUR", "USD", 1.09),
        ("GBP", "UAH", 52.10),
        ("GBP", "USD", 1.26),
        ("GBP", "EUR", 1.17),
        ("UAH", "GBP", 0.019),
        ("USD", "GBP", 0.79),
        ("EUR", "GBP", 0.85),
    ]
    
    for i, (c1, c2, rate) in enumerate(data, start=2):
        sheet[f'A{i}'] = c1
        sheet[f'B{i}'] = c2
        sheet[f'C{i}'] = rate
    
    workbook.save(filepath)
    print(f"Created default exchange rates file: {filepath}")


def generate_html(currencies, result=None, amount="", from_curr="", to_curr="", error=None):
    """Generate the HTML page for the currency converter."""
    
    # Generate currency options
    from_options = ""
    to_options = ""
    for curr in currencies:
        from_selected = " selected" if curr == from_curr else ""
        to_selected = " selected" if curr == to_curr else ""
        from_options += f'<option value="{curr}"{from_selected}>{curr}</option>\n'
        to_options += f'<option value="{curr}"{to_selected}>{curr}</option>\n'
    
    # Result message
    result_html = ""
    if error:
        result_html = f'<p style="color: red; margin-top: 15px;"><b>Помилка:</b> {error}</p>'
    elif result is not None:
        result_html = f'<p style="color: green; margin-top: 15px; font-size: 18px;"><b>{amount} {from_curr} = {result:.2f} {to_curr}</b></p>'
    
    html = f'''<!DOCTYPE html>
<html lang="uk">
<head>
    <meta charset="UTF-8">
    <title>Конвертер валют</title>
    <style>
        body {{
            font-family: Arial, sans-serif;
            max-width: 400px;
            margin: 50px auto;
            padding: 20px;
        }}
        h1 {{
            text-align: center;
        }}
        label {{
            display: block;
            margin-top: 15px;
            font-weight: bold;
        }}
        input, select {{
            width: 100%;
            padding: 8px;
            margin-top: 5px;
            font-size: 16px;
        }}
        button {{
            margin-top: 20px;
            width: 100%;
            padding: 10px;
            font-size: 16px;
            cursor: pointer;
        }}
    </style>
</head>
<body>
    <h1>Конвертер валют</h1>
    
    <form method="POST" action="/">
        <label for="amount">Сума:</label>
        <input type="number" id="amount" name="amount" step="0.01" min="0" 
               value="{amount}" required>
        
        <label for="from_currency">З валюти:</label>
        <select id="from_currency" name="from_currency" required>
            {from_options}
        </select>
        
        <label for="to_currency">В валюту:</label>
        <select id="to_currency" name="to_currency" required>
            {to_options}
        </select>
        
        <button type="submit">Конвертувати</button>
    </form>
    
    {result_html}
</body>
</html>'''
    
    return html


def application(environ, start_response):
    """WSGI application for currency conversion."""
    
    # Load exchange rates
    rates, currencies = load_exchange_rates()
    
    method = environ.get('REQUEST_METHOD', 'GET')
    
    result = None
    amount = ""
    from_curr = currencies[0] if currencies else ""
    to_curr = currencies[1] if len(currencies) > 1 else ""
    error = None
    
    if method == 'POST':
        try:
            # Read POST data
            content_length = int(environ.get('CONTENT_LENGTH', 0))
            post_data = environ['wsgi.input'].read(content_length).decode('utf-8')
            params = urllib.parse.parse_qs(post_data)
            
            amount = params.get('amount', [''])[0]
            from_curr = params.get('from_currency', [''])[0]
            to_curr = params.get('to_currency', [''])[0]
            
            if amount and from_curr and to_curr:
                amount_float = float(amount)
                
                if from_curr == to_curr:
                    result = amount_float
                elif (from_curr, to_curr) in rates:
                    rate = rates[(from_curr, to_curr)]
                    result = amount_float * rate
                else:
                    error = f"Курс для {from_curr} → {to_curr} не знайдено"
                    
        except ValueError:
            error = "Невірний формат суми"
        except Exception as e:
            error = f"Помилка: {str(e)}"
    
    # Generate HTML response
    html = generate_html(currencies, result, amount, from_curr, to_curr, error)
    
    # Send response
    status = '200 OK'
    response_headers = [
        ('Content-Type', 'text/html; charset=utf-8'),
        ('Content-Length', str(len(html.encode('utf-8'))))
    ]
    start_response(status, response_headers)
    
    return [html.encode('utf-8')]


def main():
    """Start the WSGI server."""
    host = 'localhost'
    port = 8080
    
    print(f"Starting Currency Converter Server...")
    print(f"Server running at http://{host}:{port}")
    print("Press Ctrl+C to stop the server")
    
    # Create and start the server
    httpd = make_server(host, port, application)
    
    try:
        httpd.serve_forever()
    except KeyboardInterrupt:
        print("\nServer stopped.")


if __name__ == '__main__':
    main()
